#   Leer hoja de calculo
#   pip install openpyxl tabulate
import openpyxl
from tabulate import tabulate

excel__dataframe = openpyxl.load_workbook(
    "/home/ed/otrapracticamas/Ejercicio-Video-71/src/personas.xlsx"
)  #   Cargo el archivo
dataframe = excel__dataframe.active  #   Busco el libro activo de la hoja de calculo
print(dataframe)

data = []
for row in range(1, dataframe.max_row):
    _row = [
        row,
    ]
    for col in dataframe.iter_cols(1, dataframe.max_column):
        _row.append(col[row].value)
    data.append(_row)
headers = ["#", "Id", "Name", "Company", "Email", "Mac Address"]
headers_align = ("center",) * 6
print(tabulate(data, headers=headers, tablefmt="fancy_grid", colalign=headers_align))
